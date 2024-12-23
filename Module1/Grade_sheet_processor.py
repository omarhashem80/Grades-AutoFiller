import os
import cv2
import numpy as np
import pytesseract
from image_processor import ImageProcessor
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def show_image(image, window_name="Image"):
    """
    Display an image using OpenCV.

    Parameters:
        image (numpy.ndarray): The image to be displayed.
        window_name (str): The name of the display window. Defaults to "Image".
    """
    cv2.imshow(window_name, image)
    cv2.waitKey(0)  # Waits for a key press to close the window
    cv2.destroyAllWindows()  # Closes all OpenCV windows


class GradeSheetOCR:
    def __init__(self):
        self.image_processor = ImageProcessor()

    def _process_code_ocr(self, code_image, first_stage=False):
        min_crop_size = (35, 100)
        processed_image = self._remove_table_lines(code_image)
        # show_image(processed_image)
        kernel = np.ones((2, 5), np.uint8)
        dilated_image = cv2.dilate(processed_image, kernel, iterations=10)
        # show_image(dilated_image)
        contours = self.image_processor.find_contours(dilated_image)
        bounding_boxes = self._convert_contours_to_bounding_boxes(contours)
        min_height = self._get_min_height_of_bounding_boxes(bounding_boxes)
        rows = self._club_bounding_boxes_into_rows(bounding_boxes, min_height, True)

        codes = []

        for row in rows:
            for bounding_box in row:
                x, y, w, h = bounding_box
                cropped_image = code_image[y : y + h, x : x + w]
                w -= 10
                if (
                    cropped_image.shape[0] < min_crop_size[0]
                    or cropped_image.shape[1] < min_crop_size[1] - 10
                ):
                    continue
                cropped_image = cv2.resize(
                    cropped_image, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC
                )
                if first_stage:
                    codes.insert(0, cropped_image)
                else:
                    result = pytesseract.image_to_string(
                        cropped_image,
                        config="-l eng+osd --psm 7 --oem 3 -c tessedit_char_whitelist=0123456789",
                    ).strip()
                    if result != "":
                        codes.insert(0, result)
        return codes

    def _process_code_training(self, code_image):
        cropped_images = self._process_code_ocr(code_image, True)
        codes = []
        for image in cropped_images:
            predicted_number = []
            image = self.image_processor.invert_image(
                self.image_processor.threshold_image(
                    self.image_processor.convert_to_grayscale(image)
                )
            )
            placeholder = image
            # show_image(placeholder)
            digits = 7
            cell_width = placeholder.shape[1] // digits
            predicted_number = []
            for i in range(digits):
                start_x = i * cell_width if i == 0 else i * cell_width - 5
                end_x = (
                    (i + 1) * cell_width + 5 if i < digits - 1 else placeholder.shape[1]
                )  # Handle last piece for uneven splits

                cell = image[:, start_x:end_x]
                # show_image(cell)
                result = self.image_processor.predict_digit(cell)
                predicted_number.append(result)
            code = "".join(map(str, predicted_number))
            codes.append(code)
        return codes

    def _process_names_ocr(self, image, isEnglish=True):
        min_crop_size = (35, 100)
        processed_image = self._remove_table_lines(image)
        processed_image = cv2.GaussianBlur(processed_image, (5, 5), 1)
        kernel = np.ones((2, 5), np.uint8)
        dilated_image = cv2.dilate(processed_image, kernel, iterations=15)
        # show_image(dilated_image)
        contours = self.image_processor.find_contours(dilated_image)
        bounding_boxes = self._convert_contours_to_bounding_boxes(contours)
        min_height = self._get_min_height_of_bounding_boxes(bounding_boxes)
        rows = self._club_bounding_boxes_into_rows(bounding_boxes, min_height, True)
        self._sort_rows_by_x_coordinate(rows)
        names = []
        if isEnglish:
            config = "-l eng+osd --psm 7 --oem 3 "
        else:
            config = "-l ara --psm 7 --oem 3"
        for row in rows:
            for bounding_box in row:
                x, y, w, h = bounding_box
                cropped_image = image[y : y + h, x : x + w]
                w -= 10

                if (
                    cropped_image.shape[0] < min_crop_size[0]
                    or cropped_image.shape[1] < min_crop_size[1] - 10
                ):
                    continue
                cropped_image = cv2.resize(
                    cropped_image, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC
                )
                if cropped_image.shape[0] < 60 or cropped_image.shape[1] < 150:
                    continue

                result = pytesseract.image_to_string(
                    cropped_image, config=config
                ).strip()
                # print(result)
                names.insert(0, result)
        return names

    def _process_digit_cell_ocr(self, image):
        height, width = image.shape[:2]
        rows, columns = 17, 1
        cell_height = height // rows
        cell_width = width // columns
        numbers = []
        image = self.image_processor.invert_image(
            self.image_processor.threshold_image(
                self.image_processor.convert_to_grayscale(image)
            )
        )
        ver = np.array([[1], [1], [1], [1], [1], [1], [1], [1], [1], [1]])
        vertical_lines_eroded = cv2.erode(image, ver, iterations=10)
        image -= vertical_lines_eroded
        for row in range(rows):
            start_y = row * cell_height + 10
            end_y = (row + 1) * cell_height - 10
            start_x = 10
            end_x = cell_width - 10

            cropped_image = image[start_y:end_y, start_x:end_x]
            cropped_image = cv2.resize(
                cropped_image, None, fx=1.75, fy=1.75, interpolation=cv2.INTER_CUBIC
            )
            cropped_image
            # show_image(cropped_image)
            result = pytesseract.image_to_string(
                cropped_image,
                config="-l osd --psm 10 --oem 3 -c tessedit_char_whitelist=0123456789 tessedit_unrej_turned_image=T tessdata_auto_in.allow=T textord_min_linesize=2.5",
            ).strip()
            numbers.append(result)
        return numbers

    def _process_digit_cell_training(self, image):
        height, width = image.shape[:2]
        image = self.image_processor.invert_image(
            self.image_processor.threshold_image(
                self.image_processor.convert_to_grayscale(image)
            )
        )
        ver = np.array([[1], [1], [1], [1], [1], [1], [1], [1], [1], [1]])
        vertical_lines_eroded = cv2.erode(image, ver, iterations=10)
        image -= vertical_lines_eroded
        print(width)
        rows, columns = 17, 1
        cell_height = height // rows
        cell_width = width // columns
        numbers = []
        for row in range(rows):
            start_y = row * cell_height + 10
            end_y = (row + 1) * cell_height - 10
            start_x = 10
            end_x = cell_width - 10

            cropped_image = image[start_y:end_y, start_x:end_x]
            cropped_image = cv2.resize(
                cropped_image, None, fx=1.75, fy=1.75, interpolation=cv2.INTER_CUBIC
            )
            cropped_image
            # show_image(cropped_image)
            result = self.image_processor.predict_digit(cropped_image)
            numbers.append(result)
        return numbers

    def _process_image_into_shapes_cells(self, shape_image):
        height, width = shape_image.shape[:2]
        rows, columns = 17, 2
        cell_height = height // rows
        cell_width = width // columns
        all_results = []
        shape_image = self.image_processor.invert_image(
            self.image_processor.threshold_image(
                self.image_processor.convert_to_grayscale(shape_image)
            )
        )
        ver = np.array([[1], [1], [1], [1], [1], [1], [1], [1], [1], [1]])
        vertical_lines_eroded = cv2.erode(shape_image, ver, iterations=10)
        shape_image -= vertical_lines_eroded
        # show_image(shape_image)
        for row in range(rows):
            row_results = []
            for col in range(columns):
                start_y = row * cell_height + 10
                end_y = (row + 1) * cell_height - 10
                start_x = col * cell_width + 10
                end_x = (col + 1) * cell_width - 10

                cell = shape_image[start_y:end_y, start_x:end_x]

                # show_image(cell)
                result = self._process_shape(self.image_processor.predict_symbol(cell))
                row_results.append(result)
            print(row_results)
            all_results.append(row_results)
        return all_results

    def _process_shape(self, s):
        if s == "box":
            return 0
        elif s == "correct":
            return 5
        elif s == "empty":
            return -1
        elif s.startswith("horizontal") and s[-1].isdigit():
            if int(s[-1]) == 1:
                return 0
            return 5 - int(s[-1])
        elif s.startswith("vertical") and s[-1].isdigit():
            return int(s[-1])
        elif s == "question":
            return -2
        else:
            return None

    def _split_image_sections(self, corrected_image):
        return {
            "code": corrected_image[:, : int(corrected_image.shape[1] * 0.1)],
            "arabic_name": corrected_image[
                :,
                int(corrected_image.shape[1] * 0.1) : int(
                    corrected_image.shape[1] * 0.34
                ),
            ],
            "english_name": corrected_image[
                :,
                int(corrected_image.shape[1] * 0.34) : int(
                    corrected_image.shape[1] * 0.7
                ),
            ],
            "grade_numerical": corrected_image[
                :,
                int(corrected_image.shape[1] * 0.7) : int(
                    corrected_image.shape[1] * 0.81
                ),
            ],
            "grade_shapes": corrected_image[:, int(corrected_image.shape[1] * 0.81) :],
        }

    def _remove_table_lines(self, image):
        grey = self.image_processor.convert_to_grayscale(image)
        thresholded_image = self.image_processor.threshold_image(grey)
        inverted_image = self.image_processor.invert_image(thresholded_image)

        # Vertical line removal
        ver = np.array([[1], [1], [1], [1], [1], [1], [1], [1], [1], [1]])
        vertical_lines_eroded = cv2.erode(inverted_image, ver, iterations=10)
        vertical_lines_eroded = cv2.dilate(vertical_lines_eroded, ver, iterations=10)

        hor = np.array([[1, 1, 1, 1, 1, 1]])
        horizontal_lines_eroded = cv2.erode(inverted_image, hor, iterations=10)
        horizontal_lines_eroded = cv2.dilate(
            horizontal_lines_eroded, hor, iterations=10
        )

        # Combine and remove
        combined_image = cv2.add(vertical_lines_eroded, horizontal_lines_eroded)
        combined_image_dilated = self.image_processor.dilate_image(combined_image)
        image_without_lines = cv2.subtract(inverted_image, combined_image_dilated)

        # Remove noise
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        eroded_image = cv2.erode(image_without_lines, kernel, iterations=1)
        return self.image_processor.dilate_image(eroded_image, kernel, iterations=1)

    def _convert_contours_to_bounding_boxes(self, contours):
        bounding_boxes = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            y -= 5
            h += 10
            bounding_boxes.append((x, y, w, h))
        return bounding_boxes

    def _get_min_height_of_bounding_boxes(self, bounding_boxes):
        heights = [h for _, _, _, h in bounding_boxes]
        return np.min(heights)

    def _club_bounding_boxes_into_rows(
        self, bounding_boxes, mean_height, is_sorted=False
    ):
        half_mean_height = mean_height / 2
        rows = []
        if is_sorted:
            sorted(bounding_boxes, key=lambda x: x[1])
        current_row = [bounding_boxes[0]]
        for bounding_box in bounding_boxes[1:]:
            current_bounding_box_y = bounding_box[1]
            previous_bounding_box_y = current_row[-1][1]
            distance_between_bounding_boxes = abs(
                current_bounding_box_y - previous_bounding_box_y
            )
            if distance_between_bounding_boxes <= half_mean_height:
                current_row.append(bounding_box)
            else:
                rows.append(current_row)
                current_row = [bounding_box]
        rows.append(current_row)

        return rows

    def _sort_rows_by_x_coordinate(self, rows):
        for row in rows:
            row.sort(key=lambda x: x[0])

    def _combine_number_with_shapes(self, numbers, shapes):
        combined_results = []
        # Ensure numbers and shapes have the same length
        max_length = min(len(numbers), len(shapes))

        for i in range(max_length):
            row_numbers = numbers[i] if isinstance(numbers[i], list) else [numbers[i]]
            combined_results.append(row_numbers + shapes[i])

        return combined_results

    def process_grade_sheet(
        self,
        image_path,
        code_ocr_method="training",
        number_ocr_method="training",
    ):
        # Process the image
        corrected_image = self.image_processor.process_image(image_path)
        corrected_image = corrected_image[int(corrected_image.shape[0] * 0.04) :, :]
        # Split image into sections
        sections = self._split_image_sections(corrected_image)
        # Process each section
        codes = (
            self._process_code_ocr(sections["code"])
            if code_ocr_method == "ocr"
            else self._process_code_training(sections["code"])
        )

        arabic_names = self._process_names_ocr(sections["arabic_name"], False)
        english_names = self._process_names_ocr(sections["english_name"])

        numbers = (
            self._process_digit_cell_ocr(sections["grade_numerical"])
            if number_ocr_method == "ocr"
            else self._process_digit_cell_training(sections["grade_numerical"])
        )

        shapes = self._process_image_into_shapes_cells(sections["grade_shapes"])

        # Combine results
        combined_numbers = self._combine_number_with_shapes(numbers, shapes)

        return codes, arabic_names, english_names, combined_numbers

    def create_excel_file(
        self, codes, names, english_names, values_list, output_filename="output"
    ):

        length = min(len(codes), len(names), len(english_names), len(values_list))
        print(f"Processing Excel {output_filename}")
        rows = []
        for i in range(length):

            row = {
                "Code": codes[i],
                "Student Name": names[i],
                "English Name": english_names[i],
            }
            for j, value in enumerate(values_list[i], start=1):
                if value == -1:
                    row[str(j)] = ""
                elif value == -2:
                    row[str(j)] = -2
                else:
                    row[str(j)] = value
            rows.append(row)

        df = pd.DataFrame(rows)

        excel_path = f"./outputs/{output_filename}.xlsx"
        df.to_excel(excel_path, index=False, engine="openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Students Data"

        for col_num, header in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=header)

        red_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )

        for row_num, row_data in enumerate(rows, start=2):
            col_num = 1
            ws.cell(row=row_num, column=col_num, value=row_data["Code"])
            ws.cell(row=row_num, column=col_num + 1, value=row_data["Student Name"])
            ws.cell(row=row_num, column=col_num + 2, value=row_data["English Name"])

            for j in range(4, len(row_data) + 4):
                cell_value = row_data.get(str(j - 3), "")
                cell = ws.cell(row=row_num, column=j)

                if cell_value == -2:
                    cell.fill = red_fill
                    cell.value = ""
                elif cell_value == "":
                    cell.value = ""
                else:
                    cell.value = cell_value

        wb.save(excel_path)
        print(f"Excel file '{output_filename}' has been created successfully!")
        return output_filename


if __name__ == "__main__":
    app = GradeSheetOCR()
    combinations = [("ocr", "ocr"), ("ocr", "training"), ("training", "ocr")]
    for i in range(1, 16):
        image = f"Module1/grade sheet/{i}.jpg"
        # image = cv2.imread(path)
        for j in range(3):
            codes, arabic_names, english_names, values = app.process_grade_sheet(
                image,
                code_ocr_method=combinations[j][0],
                number_ocr_method=combinations[j][1],
            )
            app.create_excel_file(
                codes,
                arabic_names,
                english_names,
                values,
                output_filename=f"output_{i}_{'Y' if combinations[j][0] == 'ocr' else 'N'}_{'Y' if combinations[j][1] == 'ocr' else 'N'}",
            )
